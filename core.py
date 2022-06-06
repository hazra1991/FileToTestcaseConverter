import csv
from collections import OrderedDict
from abc import ABC, abstractmethod
from typing import Iterable
from io import StringIO
from openpyxl import load_workbook


class Error:
    class DuplicateRowIdFound(Exception):
        """ Error class for duplicate ID """


    class DuplicateColumnFound(Exception):
        """Error class for duplicate Column"""


    class ColumnNotFound(Exception):
        """Error class for missing column name """
    
    class InvalidRecord(Exception):
        """ Error class for records not available or invalid """
    
    class PathTranslationError(Exception):
        def __init__(self,name,*a,**kw):
            msg  = f"'{name}' cannot be translated to a valid path in the provided schema file"
            super().__init__(msg,*a,**kw)


class ImmutableType(type):
    """ Base immutable type class """

    def __setattr__(self,k,val):
        msg = f'[++] class "{self.__name__}" is Immutable and doesnot support assignmnet oporation'
        raise TypeError(msg)
    def __call__(self):
        msg = f'[++]class "{self.__name__}" is Immutable and doesnot support class instantiation'
        raise TypeError(msg)
    
    def __str__(self):
        buffer = StringIO()
        buffer.write(f"{'+'*10} Global default Configuration {'+'*10}\n\n")
        for key,val in vars(self).items():
            if key == "__module__" or key == "__doc__":
                continue
            if not callable(val):
                buffer.write(f"{key} = {val}\n")
            else:
                buffer.write(f"{key} :function at {val}\n")
        return buffer.getvalue()


class ImmutableConfig(metaclass=ImmutableType):
    """ Immutable class base """


class CsvDictReader(csv.DictReader):
    """ This class will create a ordered dictionary out of the csv file """

    def __init__(self, *args, **kwargs):
        self.dict_class = kwargs.pop("dict_class", dict)
        self.ignore_header_whitespaces : bool = kwargs.pop("ignore_header_whitespaces",False)
        super().__init__(*args, **kwargs)
        # removing white spaces fromthe header columns
        if self.ignore_header_whitespaces:
            self.fieldnames = [x.strip() for x in self.fieldnames]

    def __next__(self):
        if self.line_num == 0:
            # Used only for its side effect.
            self.fieldnames
        row = next(self.reader)
        self.line_num = self.reader.line_num

        # unlike the basic reader, we prefer not to return blanks,
        # because we will typically wind up with a dict full of None
        # values
        while row == []:
            row = next(self.reader)
        
        # using the customized dict_class
        d = self.dict_class(zip(self.fieldnames, row))
        lf = len(self.fieldnames)
        lr = len(row)
        if lf < lr:
            d[self.restkey] = row[lf:]
        elif lf > lr:
            for key in self.fieldnames[lr:]:
                d[key] = self.restval
        return d


class BaseReader(ABC):
    """ Any class inherithing from this needs to return an iterable object of OrderedDict for consistancy """
        
    @abstractmethod
    def get_records(self) -> Iterable[OrderedDict]:
        """ Implement the process which reads the records fromt the give file 
        and returns an iterable or ordered Dictionary """


    @abstractmethod
    def get_headers(self) -> list:
        """ Abstract class method for all the reader classes that returns 
        the headers of the parsed input table """


class XlsxParserGen(BaseReader):
    """ Class for parsing excell file format to a dictionary """

    def __init__(self,input_file=None,trim_headers=False,**kw):
        self.input_file = input_file
        self.trim_headers = trim_headers
        self.worksheet_name = kw.pop('worksheet_name',None)
        self.wb = load_workbook(self.input_file,read_only=True, data_only=True)
        worksheet = self.wb[self.worksheet_name]
        self.values = worksheet.values      # wroksheet.values is a property generator fun hence assigning it instantiate the gen fun
        self.headers = self.__fetch_headers(worksheet)
        if self.headers is None:
            raise Error.InvalidRecord('Invalid record or worksheet is error')
 
    def get_records(self):
        # nxtRec = next(self.values)
        for nxtRec in self.values:
            if all(el is None for el in nxtRec):
                return # break can also be used for stop the gen fun
            d = OrderedDict(zip(self.headers,nxtRec))
            yield d

    def get_headers(self):
        return self.headers

    def __fetch_headers(self,worksheet):
        for i in range(worksheet.max_row):
            row = next(self.values)
            if not all(el is None for el in row):
                if self.trim_headers:
                    row = [x.strip() if x is not None else '' for x in row ]
                else:
                    row = [x if x is not None else '' for x in row ]
                return row
        return None

    def __del__(self):
        self.wb.close()


class CsvParser(BaseReader):
    """ Class for parsing the CSV file format to a dictionary  """

    def __init__(self,input_file=None,trim_headers=False,**kw):
        self.inp_file = input_file
        self.trim_headers = trim_headers
        self.fd = open(self.inp_file,'r')
        self.records = CsvDictReader(self.fd,dict_class = OrderedDict,ignore_header_whitespaces=self.trim_headers)
        self.fieldnames = self.records.fieldnames

    def get_records(self) -> Iterable[OrderedDict]:
        return self.records

    def get_headers(self):
        return self.fieldnames
    
    def __del__(self):
        self.fd.close()
